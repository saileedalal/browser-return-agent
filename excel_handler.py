"""
excel_handler.py
-----------------
All Excel reading/writing lives here (using openpyxl).

Design:
- The workbook is opened once and kept open for the whole run.
- Every time we update a row, we save the workbook immediately to disk.
  This matches the product requirement: "Immediately update Excel.
  Never wait until all rows finish."
"""

from datetime import datetime
import openpyxl
import config
from logger import get_logger

logger = get_logger()


class ExcelHandler:
    def __init__(self, file_path: str = config.EXCEL_FILE_PATH):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(self.file_path)
        self.sheet = self.workbook[config.SHEET_NAME]

        # Map column name -> column index, read from the header row (row 1).
        # This way the code does not depend on hardcoded column letters.
        self.headers = {}
        for cell in self.sheet[1]:
            if cell.value:
                self.headers[cell.value.strip()] = cell.column

    def _col(self, column_name: str) -> int:
        """Return the column index for a given header name."""
        return self.headers[column_name]

    def get_pending_rows(self):
        """
        Scan the sheet and return a list of dicts, one per row where
        Task Status == Pending. Each dict contains the row number and
        the raw values needed to process that row.
        """
        pending_rows = []

        for row_num in range(2, self.sheet.max_row + 1):  # row 1 is header
            status_cell = self.sheet.cell(
                row=row_num, column=self._col(config.COL_TASK_STATUS)
            )
            status_value = (status_cell.value or "").strip()

            if status_value == config.STATUS_PENDING:
                row_data = {
                    "row_num": row_num,
                    "platform": self._get_cell_value(row_num, config.COL_PLATFORM),
                    "order_id": self._get_cell_value(row_num, config.COL_ORDER_ID),
                    "sku": self._get_cell_value(row_num, config.COL_SKU),
                }
                pending_rows.append(row_data)

        return pending_rows

    def _get_cell_value(self, row_num: int, column_name: str):
        """Read a single cell value, stripped, or None if blank."""
        value = self.sheet.cell(row=row_num, column=self._col(column_name)).value
        if value is None:
            return None
        value = str(value).strip()
        return value if value != "" else None

    def update_row(
        self,
        row_num: int,
        task_status: str,
        reason: str = None,
        return_id: str = None,
        return_status: str = None,
        refund_amount=None,
        set_timestamp: bool = False,
    ):
        """
        Update a single row and save the workbook immediately.
        Only the fields passed in are updated; everything else is left as-is.
        """
        self.sheet.cell(
            row=row_num, column=self._col(config.COL_TASK_STATUS)
        ).value = task_status

        if reason is not None:
            self.sheet.cell(
                row=row_num, column=self._col(config.COL_REASON)
            ).value = reason

        if return_id is not None:
            self.sheet.cell(
                row=row_num, column=self._col(config.COL_RETURN_ID)
            ).value = return_id

        if return_status is not None:
            self.sheet.cell(
                row=row_num, column=self._col(config.COL_RETURN_STATUS)
            ).value = return_status

        if refund_amount is not None:
            self.sheet.cell(
                row=row_num, column=self._col(config.COL_REFUND_AMOUNT)
            ).value = refund_amount

        if set_timestamp:
            self.sheet.cell(
                row=row_num, column=self._col(config.COL_TIMESTAMP)
            ).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Save immediately - never batch updates until the end.
        self.workbook.save(self.file_path)
        logger.info(f"Updated Excel row {row_num} -> Task Status = {task_status}")

    def mark_human_review(self, row_num: int, reason: str):
        """Shortcut used whenever the agent must stop working on a row."""
        self.update_row(
            row_num,
            task_status=config.STATUS_HUMAN_REVIEW,
            reason=reason,
            set_timestamp=True,
        )
        logger.info(f"Row {row_num} marked as Human Review. Reason: {reason}")
